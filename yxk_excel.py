# -*- coding: utf-8 -*-
"""Excel 导入导出与模板。"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from yxk_account_parse import split_credential_field
from yxk_core import AccountRow, LoginStatus, QueryResult
from yxk_paths import get_base_dir

BASE_DIR = get_base_dir()
TEMPLATE_DIR = BASE_DIR / "template"
TEMPLATE_FILE = TEMPLATE_DIR / "账号导入模板.xlsx"

INPUT_HEADERS = ["账号密码", "备注"]
LEGACY_INPUT_HEADERS = ["账号", "密码", "备注"]
OUTPUT_HEADERS = [
    "问题标记",
    "状态",
    "账号",
    "密码",
    "姓名",
    "身份证号码",
    "单位名称",
    "专业职称",
    "职称代码",
    "错误信息",
    "登录方式",
    "重试次数",
    "备注",
    "原Excel行号",
]

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(color="FFFFFF", bold=True)
PASSWORD_FILL = PatternFill("solid", fgColor="FFC7CE")
PASSWORD_FONT = Font(color="9C0006", bold=True)
PROBLEM_FILL = PatternFill("solid", fgColor="FFE699")
SUCCESS_FILL = PatternFill("solid", fgColor="E2EFDA")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _template_is_current() -> bool:
    if not TEMPLATE_FILE.exists():
        return False
    try:
        wb = load_workbook(TEMPLATE_FILE, read_only=True, data_only=True)
        ws = wb["账号列表"] if "账号列表" in wb.sheetnames else wb.active
        header = _normalize_header(ws.cell(row=1, column=1).value)
        wb.close()
        return header == "账号密码"
    except Exception:
        return False


def ensure_template() -> Path:
    TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
    if _template_is_current():
        return TEMPLATE_FILE

    wb = Workbook()
    ws = wb.active
    ws.title = "账号列表"

    for col, header in enumerate(INPUT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    samples = [
        ("13982082863 Ck636489", "示例-空格分隔"),
        ("15982046865,lty623851", "示例-英文逗号分隔"),
        ("账号2345，密码jfe235", "示例-带标签逗号分隔"),
        ("账号：2345密码ddd123", "示例-带标签连续填写"),
    ]
    for row_idx, row in enumerate(samples, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if col_idx == 1:
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 28

    guide = wb.create_sheet("填写说明")
    guide["A1"] = "易学酷账号批量查询 - 输入模板说明"
    guide["A1"].font = Font(bold=True, size=14)
    lines = [
        "1. 请在“账号列表”工作表中填写数据，从第2行开始。",
        "2. “账号密码”列合并填写账号与密码，程序会自动识别拆分。",
        "3. 支持空格、英文/中文逗号、分号、冒号、竖线、换行等作为分隔符。",
        "4. 支持带标签写法，如“账号2345，密码jfe235”“账号：2345密码ddd123”。",
        "5. 若识别顺序有误，登录失败时会自动尝试账号密码互换、符号纠正等组合。",
        "6. “备注”可选，会原样带入导出结果。",
        "7. 仍兼容旧版两列模板（账号 + 密码）。",
        "8. 导出结果保存在输入 Excel 同目录下；问题账号排在最前并标色。",
        "9. 标记“⚠ 账号密码有问题”表示全部识别组合均登录失败(10013)。",
    ]
    for idx, line in enumerate(lines, start=3):
        guide[f"A{idx}"] = line
    guide.column_dimensions["A"].width = 90

    wb.save(TEMPLATE_FILE)
    return TEMPLATE_FILE


def _normalize_header(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return text.replace("*", "")


def read_accounts(excel_path: str | Path) -> list[AccountRow]:
    path = Path(excel_path)
    wb = load_workbook(path, data_only=True)
    ws = wb["账号列表"] if "账号列表" in wb.sheetnames else wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_map = {_normalize_header(value): idx for idx, value in enumerate(header_row)}

    def col(name: str) -> int | None:
        for key, idx in header_map.items():
            if key == name:
                return idx
        return None

    credential_idx = col("账号密码")
    username_idx = col("账号")
    password_idx = col("密码")
    remark_idx = col("备注")

    use_merged = credential_idx is not None
    use_legacy = username_idx is not None and password_idx is not None
    if not use_merged and not use_legacy:
        raise ValueError("Excel 缺少必填列：账号密码（或旧版 账号 + 密码）")

    accounts: list[AccountRow] = []
    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        remark = ""
        if remark_idx is not None and remark_idx < len(row) and row[remark_idx] is not None:
            remark = str(row[remark_idx]).strip()

        if use_merged:
            raw_value = row[credential_idx] if credential_idx < len(row) else None
            raw_text = "" if raw_value is None else str(raw_value).strip()
            if not raw_text:
                continue
            username_text, password_text = split_credential_field(raw_text)
            accounts.append(
                AccountRow(
                    username=username_text,
                    password=password_text,
                    remark=remark,
                    row_no=row_no,
                    raw_input=raw_text,
                )
            )
            continue

        username = row[username_idx] if username_idx < len(row) else None
        password = row[password_idx] if password_idx < len(row) else None
        username_text = "" if username is None else str(username).strip()
        password_text = "" if password is None else str(password).strip()
        if not username_text and not password_text:
            continue

        accounts.append(
            AccountRow(
                username=username_text,
                password=password_text,
                remark=remark,
                row_no=row_no,
            )
        )
    return accounts


def _style_row(ws, row_idx: int, result: QueryResult) -> None:
    if result.status == LoginStatus.PASSWORD_ERROR:
        fill = PASSWORD_FILL
        font = PASSWORD_FONT
    elif result.is_problem:
        fill = PROBLEM_FILL
        font = Font(bold=True, color="7F6000")
    else:
        fill = SUCCESS_FILL
        font = Font(color="006100")

    for col_idx in range(1, len(OUTPUT_HEADERS) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = fill
        cell.font = font
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def export_results(results: Iterable[QueryResult], output_dir: str | Path | None = None) -> Path:
    output_root = Path(output_dir or BASE_DIR / "output")
    output_root.mkdir(parents=True, exist_ok=True)

    sorted_results = sorted(list(results), key=lambda item: item.sort_key)
    problem_count = sum(1 for item in sorted_results if item.is_problem)
    password_count = sum(1 for item in sorted_results if item.status == LoginStatus.PASSWORD_ERROR)
    success_count = sum(1 for item in sorted_results if item.status == LoginStatus.SUCCESS)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_root / f"易学酷查询结果_{timestamp}.xlsx"

    wb = Workbook()
    summary = wb.active
    summary.title = "汇总"
    summary["A1"] = "易学酷批量查询结果汇总"
    summary["A1"].font = Font(bold=True, size=14)
    summary["A3"] = "总账号数"
    summary["B3"] = len(sorted_results)
    summary["A4"] = "成功"
    summary["B4"] = success_count
    summary["A5"] = "有问题账号"
    summary["B5"] = problem_count
    summary["A6"] = "其中账号密码错误"
    summary["B6"] = password_count
    summary["A8"] = "说明"
    summary["B8"] = "详细结果见“查询结果”工作表；问题账号已排在最前并标色。"

    ws = wb.create_sheet("查询结果", 0)
    for col, header in enumerate(OUTPUT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for row_idx, result in enumerate(sorted_results, start=2):
        values = [
            result.problem_label,
            result.status.value,
            result.username,
            result.password,
            result.real_name,
            result.id_card,
            result.org_name,
            result.title_name,
            result.title_code,
            result.error_msg,
            result.login_mode,
            result.retry_count,
            result.remark,
            result.row_no,
        ]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        _style_row(ws, row_idx, result)

    widths = [16, 12, 16, 16, 10, 22, 28, 18, 10, 28, 18, 10, 16, 10]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(OUTPUT_HEADERS))}{len(sorted_results) + 1}"

    wb.save(output_path)
    return output_path
